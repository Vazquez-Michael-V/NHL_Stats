#Selenium imports for website navigation.
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

#Pandas and numpy for data tables.
import pandas as pd
import numpy as np

# Date and time imports.
import time
from datetime import datetime

# Might use csv writer in future versions.
import csv

# Openpyxl for editing excel files.
from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


PATH = "C:\Program Files (x86)\chromedriver.exe" #Directory of the Chromedriver
serv = Service(PATH)
driver = webdriver.Chrome(service=serv)

WEBSITE = "https://www.nhl.com/news/nhl-stanley-cup-champions-winners-complete-list/c-287705398"
driver.get(WEBSITE)
driver.maximize_window()
web_title = driver.title
print(WEBSITE)
print(web_title)

time.sleep(20)

champions_table = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="content-wrap"]/div[3]/div[2]/div/article[1]/div[4]/div[2]/ul[1]'))
    )

# print(champions_table.text)

# Prepare string for csv.
champions_table_str = champions_table.text
champions_table_str = champions_table_str.replace(':', ',')
champions_table_str = champions_table_str.replace(", ", ",")

# print(champions_table_str)

with open('NHL_Champions.txt', 'w') as f:
    f.write("Year,Team,Coach\n")
    f.write(champions_table_str)

time.sleep(5)

df_champs = pd.read_csv('NHL_Champions.txt')
print(df_champs)

# Create excel file for end users, and put the date as the first row.
NHL_Champions_filename = 'NHL_Champions.xlsx'
with pd.ExcelWriter(NHL_Champions_filename) as writer:
    df_champs.to_excel(writer, sheet_name='Champions', index=False)

time.sleep(5)

date = f'{datetime.now(): %m.%d.%Y %I:%M %p}'
print(date)

# Putting date as first row.
wb = load_workbook(NHL_Champions_filename)
ws = wb.worksheets[0]
ws.insert_rows(1, amount=2)
ws['A1'] = f'Data as of {date}'
wb.save(NHL_Champions_filename)





